"""
Импорт данных из реестра Минцифры и ГИСП ЕРРРП в component_catalog.db.

Минцифры: российское ПО (30K записей)
  - Колонки: [0] рег.номер, [1] название ПО, [3] класс ПО, [11] правообладатель
  - Заголовки в строке 5, данные с строки 6

ГИСП ЕРРРП: российская радиоэлектронная продукция
  - Колонки: [1] предприятие, [7] реестровый номер, [12] наименование продукции, [13] ОКПД2
  - Заголовки в строке 3, данные с строки 4
"""

import os
import re
import sqlite3
import sys

DB_PATH = os.path.join(os.path.dirname(__file__), "..", "database", "component_catalog.db")


def classify_software(class_str, name):
    """Определяет категорию ПО по классу Минцифры."""
    cl = (class_str or "").lower()
    nm = (name or "").lower()

    if any(x in cl for x in ["операционн", "02.01", "02.02"]):
        return "ОС Linux", "Операционные системы"
    if any(x in cl for x in ["антивирус", "защит", "безопасност", "03.01", "03.02", "03.03", "03.04"]):
        return "СЗИ / Антивирусы", "Средства защиты"
    if any(x in cl for x in ["субд", "базы данных", "база данных", "04.05"]):
        return "СУБД", "Базы данных"
    if any(x in cl for x in ["офис", "редактор", "текстов", "электронн", "таблиц"]):
        return "Офисное ПО", "Офисные пакеты"
    if any(x in cl for x in ["браузер", "web-обозреват"]):
        return "Браузеры", ""
    if any(x in cl for x in ["виртуализ", "гипервизор", "контейнер"]):
        return "Гипервизоры", "Виртуализация"
    if any(x in cl for x in ["почт", "коммуникац", "мессенджер"]):
        return "Прикладное ПО", "Коммуникации"
    if any(x in cl for x in ["сервер приложен", "веб-сервер", "web-сервер"]):
        return "Веб-серверы", ""
    if any(x in cl for x in ["системы управлен", "erp", "crm"]):
        return "Прикладное ПО", "Корпоративное"
    return "Прикладное ПО", ""


def classify_hardware(product_name, okpd2):
    """Определяет категорию оборудования ГИСП по ОКПД2 и названию."""
    nm = (product_name or "").lower()
    okpd = (okpd2 or "")

    if any(x in nm for x in ["коммутатор", "switch"]):
        return "Сетевое оборудование", "Коммутаторы"
    if any(x in nm for x in ["маршрутизатор", "router", "роутер"]):
        return "Сетевое оборудование", "Маршрутизаторы"
    if any(x in nm for x in ["межсетев", "firewall", "брандмауэр", "экран"]):
        return "Сетевое оборудование", "МСЭ"
    if any(x in nm for x in ["точка доступа", "wi-fi", "wifi", "беспровод"]):
        return "Сетевое оборудование", "Wi-Fi"
    if any(x in nm for x in ["сервер", "server"]):
        return "Серверы", ""
    if any(x in nm for x in ["компьютер", "арм", "моноблок", "рабоч", "терминал", "тонк"]):
        return "АРМ", ""
    if any(x in nm for x in ["ноутбук", "laptop", "портативн"]):
        return "АРМ", "Ноутбуки"
    if any(x in nm for x in ["планшет", "tablet"]):
        return "АРМ", "Планшеты"
    if any(x in nm for x in ["принтер", "мфу", "печат"]):
        return "Принтеры", ""
    if any(x in nm for x in ["монитор", "дисплей"]):
        return "Мониторы", ""
    if any(x in nm for x in ["процессор", "cpu", "цпу"]):
        return "Процессоры", ""
    if any(x in nm for x in ["накопител", "ssd", "hdd", "диск"]):
        return "Накопители", ""
    if any(x in nm for x in ["модем", "modem"]):
        return "Сетевое оборудование", "Модемы"
    if any(x in nm for x in ["телефон", "ip-телефон"]):
        return "Сетевое оборудование", "IP-телефоны"
    if any(x in nm for x in ["камер", "видеонаблюд"]):
        return "Оборудование", "IP-камеры"
    if okpd.startswith("26.20"):
        return "АРМ", ""
    if okpd.startswith("26.30"):
        return "Сетевое оборудование", ""
    return "Оборудование", ""


def import_mintsifry(conn, filepath):
    """Импортирует реестр Минцифры."""
    import openpyxl

    print(f"[INFO] Импорт реестра Минцифры: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    cursor = conn.cursor()
    count = 0
    batch = []

    for row in ws.iter_rows(min_row=6, values_only=True):
        reg_num = str(row[0] or "").strip()
        name = str(row[1] or "").strip()
        sw_class = str(row[3] or "").strip()
        owner = str(row[11] or "").strip()
        owner_short = str(row[12] or "").strip()

        if not name or not reg_num:
            continue

        category, subcategory = classify_software(sw_class, name)
        vendor = owner_short if owner_short else owner
        # Упрощаем vendor
        vendor = vendor.replace('"', '').replace('«', '').replace('»', '')
        for prefix in ['АО ', 'ООО ', 'ПАО ', 'ЗАО ', 'ФГУП ']:
            if vendor.startswith(prefix):
                vendor = vendor[len(prefix):]

        product = re.sub(r'[^\w\s\-\.]', '', name.lower().replace(' ', '_'))[:100]

        batch.append((
            name, vendor, product, "", "a", category, subcategory, "mintsifry", reg_num
        ))

        if len(batch) >= 5000:
            cursor.executemany("""
                INSERT OR IGNORE INTO russian_components
                (title, vendor, product, version, part, category, subcategory, registry, registry_id)
                VALUES (?,?,?,?,?,?,?,?,?)
            """, batch)
            conn.commit()
            count += len(batch)
            batch.clear()
            print(f"  {count:,} записей...")

    if batch:
        cursor.executemany("""
            INSERT OR IGNORE INTO russian_components
            (title, vendor, product, version, part, category, subcategory, registry, registry_id)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, batch)
        conn.commit()
        count += len(batch)

    wb.close()
    print(f"[OK] Минцифры: {count:,} записей ПО")
    return count


def import_gisp(conn, filepath):
    """Импортирует ГИСП ЕРРРП."""
    import openpyxl

    print(f"[INFO] Импорт ГИСП ЕРРРП: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    cursor = conn.cursor()
    count = 0
    batch = []

    for row in ws.iter_rows(min_row=4, values_only=True):
        enterprise = str(row[1] or "").strip() if len(row) > 1 else ""
        reg_num = str(row[7] or "").strip() if len(row) > 7 else ""
        product_name = str(row[12] or "").strip() if len(row) > 12 else ""
        okpd2 = str(row[13] or "").strip() if len(row) > 13 else ""

        if not product_name:
            continue

        category, subcategory = classify_hardware(product_name, okpd2)
        vendor = enterprise.replace('"', '').replace('«', '').replace('»', '')
        for prefix in ['АКЦИОНЕРНОЕ ОБЩЕСТВО ', 'ОБЩЕСТВО С ОГРАНИЧЕННОЙ ОТВЕТСТВЕННОСТЬЮ ',
                        'ПУБЛИЧНОЕ АКЦИОНЕРНОЕ ОБЩЕСТВО ', 'АО ', 'ООО ', 'ПАО ', 'ЗАО ']:
            if vendor.upper().startswith(prefix):
                vendor = vendor[len(prefix):]

        product = re.sub(r'[^\w\s\-\.]', '', product_name.lower().replace(' ', '_'))[:100]

        batch.append((
            product_name, vendor, product, "", "h", category, subcategory, "gisp", reg_num
        ))

        if len(batch) >= 5000:
            cursor.executemany("""
                INSERT OR IGNORE INTO russian_components
                (title, vendor, product, version, part, category, subcategory, registry, registry_id)
                VALUES (?,?,?,?,?,?,?,?,?)
            """, batch)
            conn.commit()
            count += len(batch)
            batch.clear()
            print(f"  {count:,} записей...")

    if batch:
        cursor.executemany("""
            INSERT OR IGNORE INTO russian_components
            (title, vendor, product, version, part, category, subcategory, registry, registry_id)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, batch)
        conn.commit()
        count += len(batch)

    wb.close()
    print(f"[OK] ГИСП: {count:,} записей оборудования")
    return count


def main():
    conn = sqlite3.connect(DB_PATH)

    scripts_dir = os.path.dirname(__file__)

    # Минцифры
    mc_file = os.path.join(scripts_dir,
        "Экспорт Реестра (Реестр российского ПО) от 2026_04_17-00-02.xlsx")
    if os.path.exists(mc_file):
        import_mintsifry(conn, mc_file)
    else:
        print(f"[SKIP] Файл Минцифры не найден: {mc_file}")

    # ГИСП
    gisp_file = os.path.join(scripts_dir, "production-rep_res_valid_only.xlsx")
    if os.path.exists(gisp_file):
        import_gisp(conn, gisp_file)
    else:
        print(f"[SKIP] Файл ГИСП не найден: {gisp_file}")

    # Статистика
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM russian_components")
    total = c.fetchone()[0]
    c.execute("SELECT registry, COUNT(*) FROM russian_components GROUP BY registry")
    by_source = c.fetchall()
    c.execute("SELECT category, COUNT(*) FROM russian_components GROUP BY category ORDER BY COUNT(*) DESC LIMIT 15")
    by_cat = c.fetchall()

    print(f"\n{'='*50}")
    print(f"  ИМПОРТ РЕЕСТРОВ ЗАВЕРШЁН")
    print(f"{'='*50}")
    print(f"  Всего российских компонентов: {total:,}")
    for src, cnt in by_source:
        print(f"    {src}: {cnt:,}")
    print(f"\n  Топ категорий:")
    for cat, cnt in by_cat:
        print(f"    {cat}: {cnt:,}")
    print(f"{'='*50}")

    conn.close()


if __name__ == "__main__":
    main()
